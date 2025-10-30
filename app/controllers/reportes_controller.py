from flask import jsonify, request, send_file
from io import BytesIO, StringIO
from datetime import datetime, timezone, timedelta
import csv

# reportlab se importa de forma diferida dentro del handler PDF

from app.services.service_reportes import generar_resumen_panel, build_dashboard


def _to_utc_naive(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(timezone.utc).replace(tzinfo=None)


def _parse_iso_utc(s: str | None) -> datetime | None:
    if not s:
        return None
    try:
        v = s.strip()
        if v.endswith("Z"):
            v = v[:-1] + "+00:00"
        dt = datetime.fromisoformat(v)
        return _to_utc_naive(dt if dt.tzinfo else dt)
    except Exception:
        return None


def _month_range_utc(now: datetime | None = None) -> tuple[datetime, datetime]:
    now = _to_utc_naive(now or datetime.utcnow())
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    if start.month == 12:
        next_month = start.replace(year=start.year + 1, month=1)
    else:
        next_month = start.replace(month=start.month + 1)
    end = next_month - timedelta(microseconds=1)
    return start, end


def get_dashboard():
    try:
        start_iso = request.args.get("from") or request.args.get("start")
        end_iso = request.args.get("to") or request.args.get("end")
        data = generar_resumen_panel(start_iso=start_iso, end_iso=end_iso)
        return jsonify(data), 200
    except Exception as e:
        # Respuesta simple en caso de error inesperado
        return jsonify({
            "cards": {
                "pacientes_activos": {"value": 0, "suffix": "gestantes"},
                "citas_cumplidas": {"value": 0, "suffix": "%", "precision": 0},
                "alertas_generadas": {"value": 0, "suffix": "en seguimiento"},
            },
            "indicadores": {
                "altas": {"percent": 0},
                "medias": {"percent": 0},
                "alertas": {"percent": 0},
            },
            "error": str(e),
        }), 500


def get_dashboard_pdf():
    # Importar reportlab de forma diferida para no exigir la dependencia si no se usa
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Table, TableStyle
    # Preparar rango de fechas para subtítulo y cálculo
    start_iso = request.args.get("from") or request.args.get("start")
    end_iso = request.args.get("to") or request.args.get("end")
    start_dt = _parse_iso_utc(start_iso)
    end_dt = _parse_iso_utc(end_iso)
    if start_dt is None or end_dt is None:
        start_dt, end_dt = _month_range_utc()

    # Obtener datos (misma lógica del dashboard)
    data = build_dashboard(start_iso=start_dt.isoformat() + "Z", end_iso=end_dt.isoformat() + "Z")

    # Crear PDF en memoria
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 2 * cm
    y = height - margin

    # Encabezado
    logo_h = 45  # altura aprox. del logo en puntos
    try:
        logo_path = "app/Images/Ministerio de Salud.jpg"
        logo = ImageReader(logo_path)
        # Ajuste de tamaño del logo (aprox. 120x45 puntos) y alineación a la izquierda
        logo_w = 120
        c.drawImage(logo, margin, y - logo_h, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    # Espacio pequeño después del logo
    y = y - (logo_h + 12)

    c.setFillColor(colors.HexColor("#01579B"))  # azul institucional
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(width / 2, y, "Centro De Salud Jose Rubi")
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(width / 2, y - 0.7 * cm, "Reporte de Indicadores del Sistema Prenatal")

    # Subtítulo con fechas
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    rango_txt = f"Rango: {start_dt.strftime('%Y-%m-%d')}   {end_dt.strftime('%Y-%m-%d')}"
    generado_txt = f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    # Normalizar separador de fechas para evitar caracteres cuadrados
    rango_txt = f"Rango: {start_dt.strftime('%Y-%m-%d')} a {end_dt.strftime('%Y-%m-%d')}"
    c.drawCentredString(width / 2, y - 1.6 * cm, rango_txt)
    c.drawCentredString(width / 2, y - 2.2 * cm, generado_txt)

    y = y - 2.8 * cm

    # Tabla de métricas principales
    c.setFont("Helvetica-Bold", 12)
    c.drawString(margin, y, "Métricas principales")
    y -= 0.6 * cm
    # Reemplazo: tablas con estilo para métricas e indicadores
    cards = data.get("cards", {})
    metrics_header = ["Pacientes activos", "Citas cumplidas", "Alertas generadas"]
    metrics_row = [
        f"{cards.get('pacientes_activos', {}).get('value', 0)}",
        f"{cards.get('citas_cumplidas', {}).get('value', 0)}%",
        f"{cards.get('alertas_generadas', {}).get('value', 0)}",
    ]
    metrics_data = [metrics_header, metrics_row]

    available_width = width - 2 * margin
    col_width = available_width / 3.0
    header_color = colors.HexColor("#003366")

    metrics_table = Table(metrics_data, colWidths=[col_width] * 3)
    metrics_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
    ]))

    w, h = metrics_table.wrapOn(c, available_width, y)
    metrics_table.drawOn(c, margin, y - h)
    y = y - h - 12

    indicadores = data.get("indicadores", {})
    risk_header = ["Altas", "Medias", "Alertas"]
    risk_row = [
        f"{indicadores.get('altas', {}).get('percent', 0)}%",
        f"{indicadores.get('medias', {}).get('percent', 0)}%",
        f"{indicadores.get('alertas', {}).get('percent', 0)}%",
    ]
    risk_data = [risk_header, risk_row]
    risk_table = Table(risk_data, colWidths=[col_width] * 3)
    risk_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
    ]))

    w2, h2 = risk_table.wrapOn(c, available_width, y)
    risk_table.drawOn(c, margin, y - h2)
    y = y - h2 - 6

    # Pie de página
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.grey)
    c.drawString(margin, margin + 1.2 * cm, "Firma o sello digital: _________________________________")
    c.setFillColor(colors.black)
    c.drawString(margin, margin + 0.6 * cm, "Sistema SIGEPREN – Ministerio de Salud de Nicaragua – Reporte generado automáticamente")

    c.showPage()
    c.save()

    buffer.seek(0)
    filename = f"reporte_dashboard_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


def get_dashboard_excel():
    # Preparar rango y datos
    start_iso = request.args.get("from") or request.args.get("start")
    end_iso = request.args.get("to") or request.args.get("end")
    start_dt = _parse_iso_utc(start_iso)
    end_dt = _parse_iso_utc(end_iso)
    if start_dt is None or end_dt is None:
        start_dt, end_dt = _month_range_utc()
    data = build_dashboard(start_iso=start_dt.isoformat() + "Z", end_iso=end_dt.isoformat() + "Z")

    # Intentar exportar como XLSX estilado; si no hay openpyxl, usar CSV mejorado
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Estilos
        header_fill = PatternFill(fill_type="solid", fgColor="FF003366")
        white_bold = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        title_font = Font(name="Calibri", bold=True, size=14)
        section_font = Font(name="Calibri", bold=True, size=12)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        row = 1
        # Título
        title_cell = ws.cell(row=row, column=1, value="Reporte de Indicadores del Sistema Prenatal")
        title_cell.font = title_font
        title_cell.alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        # Rango y generado
        ws.cell(row=row, column=1, value="Rango").font = section_font
        ws.cell(row=row, column=2, value=start_dt.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=end_dt.strftime('%Y-%m-%d'))
        row += 1
        ws.cell(row=row, column=1, value="Generado").font = section_font
        ws.cell(row=row, column=2, value=datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'))
        row += 1

        # Línea en blanco
        row += 1

        # Encabezado sección: Métricas principales
        ws.cell(row=row, column=1, value="Métricas principales").font = section_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        # Tabla de métricas
        metrics_header = ["Pacientes activos", "Citas cumplidas (%)", "Alertas generadas (en seguimiento)"]
        metrics_values = [
            f"{data['cards'].get('pacientes_activos', {}).get('value', 0)} gestantes",
            f"{data['cards'].get('citas_cumplidas', {}).get('value', 0)} %",
            f"{data['cards'].get('alertas_generadas', {}).get('value', 0)} en seguimiento",
        ]
        for col, text in enumerate(metrics_header, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = center
            cell.border = border_all
        row += 1
        for col, text in enumerate(metrics_values, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.alignment = center
            cell.border = border_all
        row += 1

        # Línea en blanco
        row += 1

        # Encabezado sección: Indicadores de riesgo
        ws.cell(row=row, column=1, value="Indicadores de riesgo").font = section_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        # Tabla de indicadores
        risk_header = ["Altas", "Medias", "Alertas"]
        ind = data.get("indicadores", {})
        risk_values_pct = [
            ind.get('altas', {}).get('percent', 0) / 100,
            ind.get('medias', {}).get('percent', 0) / 100,
            ind.get('alertas', {}).get('percent', 0) / 100,
        ]
        for col, text in enumerate(risk_header, start=1):
            cell = ws.cell(row=row, column=col, value=text)
            cell.fill = header_fill
            cell.font = white_bold
            cell.alignment = center
            cell.border = border_all
        row += 1
        for col, val in enumerate(risk_values_pct, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = "0%"
            cell.alignment = center
            cell.border = border_all
        row += 1

        # Ajuste de anchos de columna simple basado en longitud del texto
        for col in range(1, 4):
            max_len = 0
            for r in range(1, row):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                s = v if isinstance(v, str) else str(v)
                max_len = max(max_len, len(s))
            ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 50)

        # Volcar a bytes
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        filename = f"reporte_dashboard_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception:
        # Fallback CSV estructurado si no hay openpyxl
        buf_text = StringIO()
        writer = csv.writer(buf_text)
        writer.writerow(["Reporte de Indicadores del Sistema Prenatal"])
        writer.writerow(["Rango", start_dt.strftime('%Y-%m-%d'), end_dt.strftime('%Y-%m-%d')])
        writer.writerow(["Generado", datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')])
        writer.writerow([])
        # Métricas principales en tabla
        writer.writerow(["Métricas principales"])    
        writer.writerow(["Pacientes activos", "Citas cumplidas (%)", "Alertas generadas (en seguimiento)"])
        cards = data.get("cards", {})
        writer.writerow([
            f"{cards.get('pacientes_activos', {}).get('value', 0)} gestantes",
            f"{cards.get('citas_cumplidas', {}).get('value', 0)} %",
            f"{cards.get('alertas_generadas', {}).get('value', 0)} en seguimiento",
        ])
        writer.writerow([])
        # Indicadores de riesgo en tabla
        writer.writerow(["Indicadores de riesgo"]) 
        writer.writerow(["Altas", "Medias", "Alertas"]) 
        ind = data.get("indicadores", {})
        writer.writerow([
            f"{ind.get('altas', {}).get('percent', 0)} %",
            f"{ind.get('medias', {}).get('percent', 0)} %",
            f"{ind.get('alertas', {}).get('percent', 0)} %",
        ])

        payload = buf_text.getvalue().encode("utf-8-sig")
        filename = f"reporte_dashboard_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        return send_file(
            BytesIO(payload),
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=filename,
        )


def get_dashboard_json_download():
    start_iso = request.args.get("from") or request.args.get("start")
    end_iso = request.args.get("to") or request.args.get("end")
    start_dt = _parse_iso_utc(start_iso)
    end_dt = _parse_iso_utc(end_iso)
    if start_dt is None or end_dt is None:
        start_dt, end_dt = _month_range_utc()
    data = build_dashboard(start_iso=start_dt.isoformat() + "Z", end_iso=end_dt.isoformat() + "Z")

    # Enviar como descarga JSON
    from flask import json as flask_json
    payload = flask_json.dumps(data, ensure_ascii=False).encode("utf-8")
    filename = f"reporte_dashboard_{datetime.utcnow().strftime('%Y%m%d')}.json"
    return send_file(
        BytesIO(payload),
        mimetype="application/json",
        as_attachment=True,
        download_name=filename,
    )
